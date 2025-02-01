import json
from io import BytesIO
from datetime import datetime, date

from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook, Workbook

from .models import Pessoa


def calculate_age(born):
    """Calcula a idade com base na data de nascimento."""
    today = date.today()
    return today.year - born.year - ((today.month, today.day) < (born.month, born.day))


def get_valor(age):
    """Retorna o valor com base na faixa etária."""
    if age < 21:
        return 100.00
    elif 21 <= age <= 59:
        return 150.00
    else:  # age >= 60
        return 200.00


@csrf_exempt
def upload_planilha(request):
    """
    Endpoint para receber a planilha Excel (.xlsx), processar os dados aplicando as regras:
      - Marcar como inativa pessoas menores de 18 anos.
      - Apenas pessoas ativas são retornadas.
      - Cálculo do valor conforme faixa etária.
    Retorna um JSON com a lista de pessoas ativas.
    """
    if request.method == 'POST':
        excel_file = request.FILES.get('file')
        if not excel_file:
            return JsonResponse({'error': 'Nenhum arquivo foi enviado.'}, status=400)

        try:
            wb = load_workbook(filename=excel_file, data_only=True)
            ws = wb.active
        except Exception as e:
            return JsonResponse({'error': f'Erro ao ler o arquivo: {str(e)}'}, status=400)

        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        expected_columns = ['nome', 'e-mail', 'data de nascimento', 'ativo']
        header_lower = [str(col).lower() if col else '' for col in header]
        if header_lower != expected_columns:
            return JsonResponse({
                'error': 'Estrutura da planilha inválida. As colunas devem ser: nome, e-mail, data de nascimento, ativo.'
            }, status=400)

        result = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            nome, email, data_nascimento, ativo = row

            if isinstance(data_nascimento, datetime):
                data_nascimento = data_nascimento.date()
            elif isinstance(data_nascimento, str):
                try:
                    data_nascimento = datetime.strptime(data_nascimento, '%Y-%m-%d').date()
                except Exception:
                    continue

            age = calculate_age(data_nascimento)

            if age < 18:
                ativo = False
            else:
                ativo = True

            if ativo:
                valor = get_valor(age)
                pessoa, created = Pessoa.objects.update_or_create(
                    email=email,
                    defaults={
                        'nome': nome,
                        'data_nascimento': data_nascimento,
                        'ativo': ativo,
                        'valor': valor
                    }
                )
                result.append({
                    'nome': nome,
                    'e-mail': email,
                    'data de nascimento': data_nascimento.strftime('%Y-%m-%d'),
                    'ativo': ativo,
                    'valor': f'R$ {valor:.2f}'
                })
            else:
                Pessoa.objects.update_or_create(
                    email=email,
                    defaults={
                        'nome': nome,
                        'data_nascimento': data_nascimento,
                        'ativo': False,
                        'valor': None
                    }
                )
        return JsonResponse(result, safe=False)
    else:
        return JsonResponse({'error': 'Método não permitido.'}, status=405)


def download_planilha(request):
    """
    Endpoint bônus: Gera uma planilha Excel (.xlsx) com as informações das pessoas ativas.
    """
    pessoas = Pessoa.objects.filter(ativo=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pessoas Ativas"

    headers = ['nome', 'e-mail', 'data de nascimento', 'ativo', 'valor']
    ws.append(headers)

    for pessoa in pessoas:
        ws.append([
            pessoa.nome,
            pessoa.email,
            pessoa.data_nascimento.strftime('%Y-%m-%d'),
            pessoa.ativo,
            f'R$ {pessoa.valor:.2f}' if pessoa.valor else ''
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=pessoas_ativas.xlsx'

    buffer = BytesIO()
    wb.save(buffer)
    response.write(buffer.getvalue())
    return response

